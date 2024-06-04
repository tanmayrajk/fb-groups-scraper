import glob
import json
import re
from post import scrape_post
import pandas as pd
from openpyxl import load_workbook
from selenium.webdriver.chrome.webdriver import WebDriver

def scrape_from_links(browser: WebDriver, batch_size: int):
    """
    Scrapes from posts in the directory ("links_*.json")
    and saves them in the directory ("posts_*.json").
    """

    pattern = 'links_*.json'
    file_list = sorted(glob.glob(pattern), key=lambda x: int(re.search(r'\d+', x).group()))
    all_data = []

    for file in file_list:
        with open(file, 'r') as file:
            data = json.load(file)
            all_data.extend(data)

    post_count = 0
    posts = []

    for link in all_data:
        data = scrape_post(browser, link)
        posts.append(data)
        post_count += 1

        if post_count % batch_size == 0:
            print(f"Saving batch of {batch_size} posts...")
            with open(f"posts_{post_count}.json", "w") as file:
                json.dump(posts, file, indent=4)
            posts.clear()

    if len(posts) > 0:
        print(f"Saving final batch of posts...")
        with open(f"posts_{post_count}.json", "w") as file:
            json.dump(posts, file, indent=4)
        posts.clear()

def posts_to_excel(output: str):
    """
    Converts the posts_*.json files in 
    the directory to a single output.xlsx.
    """

    pattern = 'posts_*.json'
    file_list = sorted(glob.glob(pattern), key=lambda x: int(re.search(r'\d+', x).group()))
    all_data = []

    for file in file_list:
        with open(file, 'r') as file:
            data = json.load(file)
            all_data.extend(data)

    df = pd.DataFrame((all_data))
    output_file = output
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    for cell in ws['I'][1:]:
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"

    wb.save(output_file)

    print(f"Data successfully written to {output_file}")